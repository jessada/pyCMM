from pycmm.template import pyCMMBase
from openpyxl import load_workbook
#from xlrd import open_workbook
#from xlrd.sheet import ctype_text 


class XlsUtils(pyCMMBase):
    """ Encapsulated util to handle excel sheet """

    def __init__(self, xls_file):
        pyCMMBase.__init__(self)
        self.__wb = load_workbook(xls_file, read_only=True)

    @property
    def wb(self):
        return self.__wb

    @property
    def nsheets(self):
        return len(self.wb.get_sheet_names())

    def get_sheet_by_name(self, sheet_name):
        return self.wb[sheet_name]

    def get_sheet_by_idx(self, sheet_idx):
        sheet_name = self.wb.get_sheet_names()[sheet_idx]
        return self.get_sheet_by_name(sheet_name)

    def get_col_idx(self, col_name, ws=None, sheet_idx=0):
        if ws is None:
            ws = self.get_sheet_by_idx(sheet_idx)
        for col_idx in xrange(1, ws.max_column+1):
            if ws.cell(row=1, column=col_idx).value == col_name:
                return col_idx
        return -1

    def get_sheet_idx(self, sheet_name):
        sheet_names = self.wb.get_sheet_names()
        for sheet_idx in xrange(len(sheet_names)):
            if sheet_names[sheet_idx] == sheet_name:
                return sheet_idx
        return -1

    def compare_vals(self,
                     col_name,
                     row_idx1,
                     row_idx2,
                     sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        col_idx = self.get_col_idx(col_name, ws=ws)
        val1 = ws.cell(row=row_idx1, column=col_idx).value
        val2 = ws.cell(row=row_idx2, column=col_idx).value
        return val1 == val2

    def count_rows(self, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.max_row

    def get_cell_rgb(self, row, col, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.cell(row=row, column=col).fill.start_color.index

    def get_cell_value(self, row, col, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.cell(row=row, column=col).value

    def count_cols(self, col_name1=None, col_name2=None, sheet_idx=0):
        """
        given param 'col_name1' and 'col_name2', the function will number of
        columns from col_name1 to col_name2.
        Otherwise, it will count all the columns
        """
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        if col_name1 is not None:
            col_idx1 = self.get_col_idx(col_name1, ws=ws)
            col_idx2 = self.get_col_idx(col_name2, ws=ws)
            return col_idx2 - col_idx1 + 1
        return ws.max_column

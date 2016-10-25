from collections import OrderedDict
from xlsxwriter.worksheet import Worksheet
from xlsxwriter import Workbook
from openpyxl import load_workbook
from pycmm.template import pyCMMBase
from pycmm.cmmlib.colorlib import COLORS_RGB

NO_COLOR = 'no_color'


class CMMWorksheet(Worksheet, pyCMMBase):
    """ a customized inherited from xlsxwriter.worksheet.Worksheet """

    def __init__(self, 
                 dflt_fmt=None,
                 **kwargs
                 ):
        self.__dflt_fmt = dflt_fmt
        super(CMMWorksheet, self).__init__(**kwargs)
        pyCMMBase.__init__(self)

    def write(self, row, col, val, cell_format=None):
        if (cell_format is None) and (self.__dflt_fmt is not None):
            cell_format = self.__dflt_fmt
        super(CMMWorksheet, self).write(row, col, val, cell_format) 

class CMMWorkbook(Workbook, pyCMMBase):
    """ a customized workbook inherited from xlsxwriter.Workbook """

    def __init__(self, 
                 **kwargs
                 ):
        super(CMMWorkbook, self).__init__(**kwargs)
        pyCMMBase.__init__(self)
        self.__dflt_hash_fmt = {'font_name': 'Arial', 'font_size': 9}
        self.__dflt_fmt = self.add_format(self.__dflt_hash_fmt)

    def get_raw_repr(self):
        raw_repr = OrderedDict()
        raw_repr["default cell format"] = self.__dflt_hash_fmt
        return raw_repr

    def _add_sheet(self, name, is_chartsheet):
        # Utility for shared code in add_worksheet() and add_chartsheet().
        sheet_index = len(self.worksheets_objs)
        name = self._check_sheetname(name, is_chartsheet)
        # Initialisation data to pass to the worksheet.
        init_data = {
            'name': name,
            'index': sheet_index,
            'str_table': self.str_table,
            'worksheet_meta': self.worksheet_meta,
            'optimization': self.optimization,
            'tmpdir': self.tmpdir,
            'date_1904': self.date_1904,
            'strings_to_numbers': self.strings_to_numbers,
            'strings_to_formulas': self.strings_to_formulas,
            'strings_to_urls': self.strings_to_urls,
            'default_date_format': self.default_date_format,
            'default_url_format': self.default_url_format,
            'excel2003_style': self.excel2003_style,
        }
        if is_chartsheet:
            worksheet = Chartsheet()
        else:
            worksheet = CMMWorksheet(dflt_fmt=self.__dflt_fmt)
        worksheet._initialize(init_data)
        self.worksheets_objs.append(worksheet)
        self.sheetnames.append(name)
        return worksheet

    def add_colors_format(self, added_hash_fmt):
        fmts = OrderedDict()
        base_hash_fmt = self.__dflt_hash_fmt.copy()
        for key in added_hash_fmt:
            base_hash_fmt[key] = added_hash_fmt[key]
        fmts[NO_COLOR] = self.add_format(base_hash_fmt)
        for color in COLORS_RGB:
            color_hash_fmt = base_hash_fmt.copy()
            color_hash_fmt['bg_color'] = COLORS_RGB[color]
            fmts[color] = self.add_format(color_hash_fmt)
        return fmts

def gen_colors_wb():
    wb = CMMWorkbook(filename="colors_book.xlsx")
    ws = wb.add_worksheet("color_sheet")
    colors_fmt = wb.add_colors_format({})
    row_idx = 0
    for color in COLORS_RGB:
        ws.write(row_idx, 0, color, cell_format=colors_fmt[color])
        row_idx += 1
    wb.close()

class XlsUtils(pyCMMBase):
    """ Encapsulated util to handle excel sheet """

    def __init__(self, xls_file):
        pyCMMBase.__init__(self)
        self.__wb = load_workbook(xls_file, read_only=True)

    @property
    def wb(self):
        return self.__wb

    @property
    def nsheets(self):
        return len(self.wb.get_sheet_names())

    def get_sheet_by_name(self, sheet_name):
        return self.wb[sheet_name]

    def get_sheet_by_idx(self, sheet_idx):
        sheet_name = self.wb.get_sheet_names()[sheet_idx]
        return self.get_sheet_by_name(sheet_name)

    def get_col_idx(self, col_name, ws=None, sheet_idx=0):
        if ws is None:
            ws = self.get_sheet_by_idx(sheet_idx)
        for col_idx in xrange(1, ws.max_column+1):
            if ws.cell(row=1, column=col_idx).value == col_name:
                return col_idx
        return -1

    def get_sheet_idx(self, sheet_name):
        sheet_names = self.wb.get_sheet_names()
        for sheet_idx in xrange(len(sheet_names)):
            if sheet_names[sheet_idx] == sheet_name:
                return sheet_idx
        return -1

    def compare_vals(self,
                     col_name,
                     row_idx1,
                     row_idx2,
                     sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        col_idx = self.get_col_idx(col_name, ws=ws)
        val1 = ws.cell(row=row_idx1, column=col_idx).value
        val2 = ws.cell(row=row_idx2, column=col_idx).value
        return val1 == val2

    def count_rows(self, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.max_row

    def get_cell_rgb(self, row, col, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.cell(row=row, column=col).fill.start_color.index

    def get_cell_value(self, row, col, sheet_idx=0):
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        return ws.cell(row=row, column=col).value

    def count_cols(self, col_name1=None, col_name2=None, sheet_idx=0):
        """
        given param 'col_name1' and 'col_name2', the function will number of
        columns from col_name1 to col_name2.
        Otherwise, it will count all the columns
        """
        ws = self.get_sheet_by_idx(sheet_idx=sheet_idx)
        if col_name1 is not None:
            col_idx1 = self.get_col_idx(col_name1, ws=ws)
            col_idx2 = self.get_col_idx(col_name2, ws=ws)
            return col_idx2 - col_idx1 + 1
        return ws.max_column
